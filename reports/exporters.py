"""
CSV and Excel export utilities.

Usage:
    from reports.exporters import export_trades_csv, export_trades_excel
    response = export_trades_csv(positions_queryset)  # returns HttpResponse
"""
import csv
import io
from datetime import date

from django.http import HttpResponse


def export_trades_csv(positions_qs, filename: str = 'trades.csv') -> HttpResponse:
    """Export a Position queryset as a CSV download."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Symbol', 'Strategy', 'Side', 'Entry Price', 'Exit Price',
        'Quantity', 'Stop Loss', 'Target', 'P&L', 'Status',
        'Opened At', 'Closed At',
    ])

    for pos in positions_qs:
        writer.writerow([
            pos.id,
            pos.symbol.symbol,
            pos.strategy,
            pos.side,
            pos.entry_price,
            pos.exit_price or '',
            pos.quantity,
            pos.stop_loss,
            pos.target,
            pos.pnl or '',
            pos.status,
            pos.opened_at.strftime('%Y-%m-%d %H:%M:%S') if pos.opened_at else '',
            pos.closed_at.strftime('%Y-%m-%d %H:%M:%S') if pos.closed_at else '',
        ])

    return response


def export_trades_excel(positions_qs, filename: str = 'trades.xlsx') -> HttpResponse:
    """Export a Position queryset as an Excel (.xlsx) download."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    headers = [
        'ID', 'Symbol', 'Strategy', 'Side', 'Entry Price', 'Exit Price',
        'Quantity', 'Stop Loss', 'Target', 'P&L', 'Status',
        'Opened At', 'Closed At',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    for row_num, pos in enumerate(positions_qs, 2):
        pnl_val = float(pos.pnl) if pos.pnl is not None else None
        ws.append([
            pos.id,
            pos.symbol.symbol,
            pos.strategy,
            pos.side,
            float(pos.entry_price),
            float(pos.exit_price) if pos.exit_price else '',
            pos.quantity,
            float(pos.stop_loss),
            float(pos.target),
            pnl_val if pnl_val is not None else '',
            pos.status,
            pos.opened_at.strftime('%Y-%m-%d %H:%M:%S') if pos.opened_at else '',
            pos.closed_at.strftime('%Y-%m-%d %H:%M:%S') if pos.closed_at else '',
        ])
        # Colour P&L cell
        pnl_cell = ws.cell(row=row_num, column=10)
        if pnl_val is not None:
            if pnl_val >= 0:
                pnl_cell.fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
            else:
                pnl_cell.fill = PatternFill(fill_type='solid', fgColor='FFC7CE')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
